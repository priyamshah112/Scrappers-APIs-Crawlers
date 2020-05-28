''' Code By Priyam Shah 
Contact - priyamshah112@gmail.com, +91 7738478888
Solution - Scrape Code for Career Site scraping based on company names

Input Dir -  For uploading all input xlsx files

Output Dir - For downloading all output xlsx files with scraped data
'''

import openpyxl #Lib for excel files
import requests #Website request library
from bs4 import BeautifulSoup #Scraping Library
import lxml
import re
import time

print("Code started")
my_path = "Input/Priyam.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
print(my_sheet_obj.max_row)
for i in range(2,my_sheet_obj.max_row+1):
  #Company name
  try:
    cname = my_sheet_obj.cell(row=i,column=2).value  
    print(cname)
    if cname:
      #check website link
      flag = False

      website_link = my_sheet_obj.cell(row=i,column=3).value
      print(website_link)
      if website_link:
        #check careers page or jobs page link parse site
        print("Parse and write in excel link or empty")

        time.sleep(5)
        
        try:
          source_code = requests.get(website_link)
        except:
          website_link = "http://"+website_link
          source_code = requests.get(website_link)

        plain_text = source_code.text
        soup = BeautifulSoup(plain_text,'lxml')

        search_list = ["Career","career","careers","job","jobs","Jobs","hire","hiring"]

        for link in soup.findAll('a',href=re.compile(r'\b(?:%s)\b' % '|'.join(search_list))):
          #print(link)
          href = link.get("href")
          print("-------Hooray link found------")
          ans = ""
          if ('www' or 'http') in href:
            ans = href

            print("pre check ",ans,str(my_sheet_obj.cell(row=i,column=3).value))
            
            if str(ans) != str(my_sheet_obj.cell(row=i,column=3).value):

              my_sheet_obj.cell(row=i,column=4).value = str(ans)
              my_wb_obj.save(my_path)
              print("Saved career site to sheet")
              print("-------------------")
              flag = True
              break

          else:
            ans = website_link+href
            
            print("pre check ",ans,str(my_sheet_obj.cell(row=i,column=3).value))

            if str(ans) != str(my_sheet_obj.cell(row=i,column=3).value):
              my_sheet_obj.cell(row=i,column=4).value = str(ans)
              my_wb_obj.save(my_path)
              print("Saved career site to sheet")
              print("-------------------")
              flag = True
              break


        if flag:
          flag = False
          continue

        else:
          #Career Site not found in website
          #Google search approach
          print("Not found career site in website...google searching now")

          headers_Get = {
                  'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:49.0) Gecko/20100101 Firefox/49.0',
                  'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                  'Accept-Language': 'en-US,en;q=0.5',
                  'Accept-Encoding': 'gzip, deflate',
                  'DNT': '1',
                  'Connection': 'keep-alive',
                  'Upgrade-Insecure-Requests': '1'
              }
          s = requests.Session()


          for j in range(len(search_list)):
            #modified cname
            tname = website_link + " " + search_list[j]

            url = 'http://www.google.com/search?q=' + tname + '&ie=utf-8&oe=utf-8'
            r = s.get(url, headers=headers_Get)

            soup = BeautifulSoup(r.text, "html.parser")
            output = []
            for searchWrapper in soup.find_all('div', {'class':'r'}):
              
              print("if ",search_list[j].lower(),searchWrapper.find('h3').text)
              if search_list[j].lower().find(searchWrapper.find('h3').text):
                url = searchWrapper.find('a')["href"]
                output.append(url)
                break

            print("Google link found")
            #print(output)
            
            print("pre check ",output[0],str(my_sheet_obj.cell(row=i,column=3).value))

            if str(output[0]) != str(my_sheet_obj.cell(row=i,column=3).value):
              my_sheet_obj.cell(row=i,column=4).value = str(output[0])
              my_wb_obj.save(my_path)
              print("Saved career site to sheet")
              print("-------------------")
              break

      else:
        flag1 = False
        #Google search company name get website link
        print("Google search name and get link")
        headers_Get = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:49.0) Gecko/20100101 Firefox/49.0',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip, deflate',
                'DNT': '1',
                'Connection': 'keep-alive',
                'Upgrade-Insecure-Requests': '1'
            }
        s = requests.Session()

        url = 'http://www.google.com/search?q=' + cname + '&ie=utf-8&oe=utf-8'
        r = s.get(url, headers=headers_Get)

        soup = BeautifulSoup(r.text, "html.parser")
        output = []
        for searchWrapper in soup.find_all('div', {'class':'r'}):
          
          print(cname.lower(),searchWrapper.find('h3').text)
          if cname.lower().find(searchWrapper.find('h3').text):
            url = searchWrapper.find('a')["href"]
            output.append(url)
            break

        print("Google link found")
        print(output)
        my_sheet_obj.cell(row=i,column=3).value = str(output[0])

        my_wb_obj.save(my_path)
        print("Saved website to sheet")
        print("-------------------")
        print("Perform parsing if company website available else pass")
        
        if len(output)>=1:
          time.sleep(5)
          
          website_link = output[0].split("/")
          print("updating url")
          #print(website_link)
          website_link = "https://"+website_link[2]
          print(website_link)
          try:
            source_code = requests.get(website_link)
          except:
            website_link = "http://"+website_link
            source_code = requests.get(website_link)
          
          
          plain_text = source_code.text
          soup = BeautifulSoup(plain_text,'lxml')

          search_list = ["Career","career","job","jobs","Jobs"]

          print("finding careers site")
          for link in soup.findAll('a',href=re.compile(r'\b(?:%s)\b' % '|'.join(search_list))):
            #print(link)
            href = link.get("href")
            print("-------Hooray link found------")

            if ('www' or 'http') in href:
              ans = href
              print("pre check ",ans,str(my_sheet_obj.cell(row=i,column=3).value))

              if str(ans) != str(my_sheet_obj.cell(row=i,column=3).value):              
                
                my_sheet_obj.cell(row=i,column=4).value = str(ans)

                my_wb_obj.save(my_path)
                print("Saved career site to sheet")
                print("-------------------")
                flag1 = True
                break

            else:
              ans = website_link+href
              #print(website_link+href)
              print("pre check ",ans,str(my_sheet_obj.cell(row=i,column=3).value))

              if str(ans) != str(my_sheet_obj.cell(row=i,column=3).value):
                my_sheet_obj.cell(row=i,column=4).value = str(ans)
                my_wb_obj.save(my_path)
                print("Saved career site to sheet")
                print("-------------------")
                flag1 = True
                break

          
          if flag1:
            flag1 = False
            continue
          else:

            #Career Site not found in website
            #Google search approach
            print("Not found career site in website...google searching now else code")
    
            search_list = ["Career","career","careers","job","jobs","Jobs","hire","hiring"]
    
            headers_Get = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:49.0) Gecko/20100101 Firefox/49.0',
                    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                    'Accept-Language': 'en-US,en;q=0.5',
                    'Accept-Encoding': 'gzip, deflate',
                    'DNT': '1',
                    'Connection': 'keep-alive',
                    'Upgrade-Insecure-Requests': '1'
                }
            s = requests.Session()


            for j in range(len(search_list)):
              #modified cname
              tname = website_link + " " + search_list[j]

              print("else code tname ",tname)

              url = 'http://www.google.com/search?q=' + tname + '&ie=utf-8&oe=utf-8'
              r = s.get(url, headers=headers_Get)

              soup = BeautifulSoup(r.text, "html.parser")
              output = []
              
              for searchWrapper in soup.find_all('div', {'class':'r'}):
                
                print("else ",search_list[j].lower(),searchWrapper.find('h3').text)
                if search_list[j].lower().find(searchWrapper.find('h3').text):
                  url = searchWrapper.find('a')["href"]
                  output.append(url)
                  break

              print("Google link found else code")
              
              #print(output)
              
              print("pre check ",output[0],str(my_sheet_obj.cell(row=i,column=3).value))

              if str(output[0]) != str(my_sheet_obj.cell(row=i,column=3).value):
                
                my_sheet_obj.cell(row=i,column=4).value = str(output[0])
                my_wb_obj.save(my_path)
                print("Saved career site to sheet")
                print("-------------------")
                break              
    
    else:
      print("No company name or Invalid name")
      pass
  


  except Exception as e:
    print(e)
    time.sleep(5)
    continue

  print("Process Completed")
